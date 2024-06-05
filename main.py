import shutil
from win32com import client
import sys
import csv
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import (QComboBox, QCheckBox, QUndoCommand, QFileDialog, QHBoxLayout, QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QMessageBox, QDateEdit, QHeaderView)
from PyQt5.QtCore import QDate, Qt
from PyQt5.QtGui import QColor, QIcon
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
import gc  

class EditDateCommand(QUndoCommand):
    def __init__(self, date_edit, old_date, new_date):
        super().__init__()
        self.date_edit = date_edit
        self.old_date = old_date
        self.new_date = new_date

    def redo(self):
        self.date_edit.setDate(self.new_date)

    def undo(self):
        self.date_edit.setDate(self.old_date)

class EditComboCommand(QUndoCommand):
    def __init__(self, check_box, old_value, new_value):
        super().__init__()
        self.check_box = check_box
        self.old_value = old_value
        self.new_value = new_value

    def redo(self):
        self.check_box.setChecked(self.new_value)

    def undo(self):
        self.check_box.setChecked(self.old_value)


class WheelIgnoredComboBox(QComboBox):
    def wheelEvent(self, event):
        event.ignore()

class WheelIgnoredDateEdit(QDateEdit):
    def wheelEvent(self, event):
        event.ignore()
        
class SchemeSelectionDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Выбрать исполнительную схему')
        self.layout = QtWidgets.QVBoxLayout(self)

        
        self.table = QtWidgets.QTableWidget(0, 3)
        self.layout.addWidget(self.table)
        self.table.setHorizontalHeaderLabels(['Наименование схемы', '№', 'Примечание'])
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)
        self.setGeometry(100, 100, 800, 600)
        self.populate_table()
        self.table.resizeColumnsToContents()

        self.ok_button = QtWidgets.QPushButton('OK')
        self.ok_button.clicked.connect(self.accept_selection)
        self.cancel_button = QtWidgets.QPushButton('Отмена')
        self.cancel_button.clicked.connect(self.reject)

        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        self.layout.addLayout(button_layout)

    def populate_table(self):
        scheme_data = self.parent().get_scheme_data()
        self.table.setRowCount(len(scheme_data))
        for row_index, (name, number, note) in enumerate(scheme_data):
            self.table.setItem(row_index, 0, QtWidgets.QTableWidgetItem(name))
            self.table.setItem(row_index, 1, QtWidgets.QTableWidgetItem(number))
            self.table.setItem(row_index, 2, QtWidgets.QTableWidgetItem(note))

    def accept_selection(self):
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, 'Ошибка', 'Не выбраны строки для добавления.')
            return

        selected_items = []
        for index in selected_rows:
            row = index.row()
            scheme = self.table.item(row, 0).text()
            number = self.table.item(row, 1).text()
            selected_items.append(f"{scheme} - {number}")

        selected_row = self.parent().table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, 'Ошибка', 'Не выбрана строка в журнале.')
            return

        existing_item = self.parent().table.item(selected_row, 13)
        existing_text = existing_item.text() if existing_item else ""
        new_text = "; ".join(selected_items)
        final_text = existing_text + "; " + new_text if existing_text else new_text
        self.parent().table.setItem(selected_row, 13, QtWidgets.QTableWidgetItem(final_text))
        self.accept()


class AgreementSelectionDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Выбрать согласование')
        self.layout = QtWidgets.QVBoxLayout(self)

        self.table = QtWidgets.QTableWidget(0, 1)  
        self.layout.addWidget(self.table)
        self.table.setHorizontalHeaderLabels(['Наименование документа'])
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)

        self.populate_table()
        self.table.resizeColumnsToContents()
        self.setGeometry(100, 100, 800, 600)
        self.ok_button = QtWidgets.QPushButton('OK')
        self.ok_button.clicked.connect(self.accept_selection)
        self.cancel_button = QtWidgets.QPushButton('Отмена')
        self.cancel_button.clicked.connect(self.reject)

        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        self.layout.addLayout(button_layout)

    def populate_table(self):
        agreement_data = self.parent().get_agreement_data()
        self.table.setRowCount(len(agreement_data))
        for row_index, (name, ) in enumerate(agreement_data):
            self.table.setItem(row_index, 0, QtWidgets.QTableWidgetItem(name))

    def accept_selection(self):
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, 'Ошибка', 'Не выбраны строки для добавления.')
            return

        selected_items = []
        for index in selected_rows:
            row = index.row()
            document = self.table.item(row, 0).text()
            selected_items.append(f"{document}")

        selected_row = self.parent().table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, 'Ошибка', 'Не выбрана строка в журнале.')
            return

        existing_item = self.parent().table.item(selected_row, 10)
        existing_text = existing_item.text() if existing_item and existing_item.text() != "Отсутствуют" else ""
        new_text = "; ".join(selected_items)
        final_text = existing_text + "; " + new_text if existing_text else new_text
        self.parent().table.setItem(selected_row, 10, QtWidgets.QTableWidgetItem(final_text))
        self.accept()


class MTRSelectionDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Выбрать МТР')
        self.layout = QtWidgets.QVBoxLayout(self)

        
        self.table = QtWidgets.QTableWidget(0, 3)  
        self.layout.addWidget(self.table)
        self.table.setHorizontalHeaderLabels(['Объект контроля', 'Сертификаты, паспорта и иные документы', 'Акты входного контроля'])
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)  
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)  

        self.setGeometry(100, 100, 800, 600)
        self.table.setStyleSheet("QTableWidget::item:selected { background-color: #add8e6; }")

        self.populate_table()

        self.table.resizeColumnsToContents()
        self.ok_button = QtWidgets.QPushButton('OK')
        self.ok_button.clicked.connect(self.accept_selection)
        self.cancel_button = QtWidgets.QPushButton('Отмена')
        self.cancel_button.clicked.connect(self.reject)
        
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        self.layout.addLayout(button_layout)
    

    def populate_table(self):
        mtr_data = self.parent().get_mtr_data()
        self.table.setRowCount(len(mtr_data))

        for row_index, row_data in enumerate(mtr_data):
            for col_index in range(0, self.table.columnCount()):
                
                if col_index < len(row_data):
                    self.table.setItem(row_index, col_index, QtWidgets.QTableWidgetItem(row_data[col_index]))
                else:
                    
                    print(f"Missing data for row {row_index} column {col_index}")

    def accept_selection(self):
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, 'Ошибка', 'Не выбраны строки для добавления.')
            return

        selected_items = []
        for index in selected_rows:
            row = index.row()
            name = self.table.item(row, 0).text()
            qty = self.table.item(row, 1).text()
            unit = self.table.item(row, 2).text()
            selected_items.append(f"{name} - {qty} {unit}")


        
        selected_row = self.parent().table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, 'Ошибка', 'Не выбрана строка в журнале.')
            return

        existing_item = self.parent().table.item(selected_row, 9)
        existing_text = existing_item.text() if existing_item and existing_item.text() != "Не применялись" else ""
        new_text = "; ".join(selected_items)
        final_text = existing_text + "; " + new_text if existing_text else new_text
        self.parent().table.setItem(selected_row, 9, QtWidgets.QTableWidgetItem(final_text))
        self.accept()


class VolumeSelectionDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Выбрать объём работы')
        self.layout = QtWidgets.QVBoxLayout(self)

        self.table = QtWidgets.QTableWidget(0, 3)  
        self.layout.addWidget(self.table)
        self.setGeometry(100, 100, 800, 600)
        self.ok_button = QtWidgets.QPushButton('OK')
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QtWidgets.QPushButton('Отмена')
        self.cancel_button.clicked.connect(self.reject)
        self.table.resizeColumnsToContents()
        button_layout = QtWidgets.QHBoxLayout()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        self.layout.addLayout(button_layout)

        self.selected_data = None

        self.table.cellDoubleClicked.connect(self.cell_double_clicked)

    def cell_double_clicked(self, row, column):
        
        num, ok = QtWidgets.QInputDialog.getInt(self, "Введите объем", "Объем работы:")
        if ok:
            
            selected_row = self.parent().table.currentRow()
            existing_data_item = self.parent().table.item(selected_row, 7)

            
            existing_data = existing_data_item.text() if existing_data_item else ""
            
            work_type = self.table.item(row, 0).text()
            volume = self.table.item(row, 1).text()
            unit = self.table.item(row, 2).text() if self.table.item(row, 2) else ""
            
            
            new_selection = f"{unit}_{work_type}_{num}_{volume}"

            
            if existing_data:
                self.selected_data = f"{existing_data}; {new_selection}".strip()
            else:
                self.selected_data = new_selection
        
            try:
                with open('docs/вор.csv', 'a+', newline='', encoding='utf-8') as file:
                    writer = csv.writer(file)
                    file.seek(0)
                    reader = csv.reader(file)

                    _, work_type, num, volume = new_selection.split("_")
                    aosr_number = self.parent().table.item(self.parent().table.currentRow(), 0).text()
                    
                    writer.writerow([work_type, volume, num, aosr_number])
            except Exception as e:
                QMessageBox.warning(self, 'Ошибка', f'Не удалось сохранить данные: {e}')
            self.accept()
        else:
            self.selected_data = None
            self.reject()  

    def set_data(self, data):
        self.table.clear()
        self.table.setRowCount(len(data))
        self.table.setColumnCount(3)
        for row_index, row_data in enumerate(data):
            for col_index, value in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(value)
                self.table.setItem(row_index, col_index, item)
        self.table.resizeColumnsToContents()
        self.table.setColumnHidden(2, True)




class NumericTableWidgetItem(QTableWidgetItem):
    def __lt__(self, other):
        try:
            return int(self.text()) < int(other.text())
        except ValueError:
            return self.text() < other.text()

class EditCellCommand(QUndoCommand):
    def __init__(self, table, row, column, old_value, new_value):
        super().__init__()
        self.table = table
        self.row = row
        self.column = column
        self.old_value = old_value
        self.new_value = new_value
        self.setText(f"Edit cell at ({row}, {column})")

    def redo(self):
        self.table.item(self.row, self.column).setText(self.new_value)

    def undo(self):
        self.table.item(self.row, self.column).setText(self.old_value)

class AOSRApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.undo_stack = QtWidgets.QUndoStack(self)
        self.row_modified = {}
        self.main_table = 0
        self.other_tables = {}
        self.current_region = 'информация_кз'
        self.initUI()

    def initUI(self):
        self.setWindowTitle('АОСР - Журнал АОСР')
        self.setGeometry(100, 100, 800, 600)
        self.tab_widget = QTabWidget(self)
        self.tab_widget.setTabPosition(QTabWidget.South)
        self.setCentralWidget(self.tab_widget)

        self.tabs = {}
        tab_names = [
            ('Журнал АОСР', 'images/2.png'),
            ('Информация', 'images/7.png'),
        ]
        for name, icon_path in tab_names:
            tab = QWidget()
            self.tabs[name] = tab
            self.tab_widget.addTab(tab, QIcon(icon_path), name)

        self.setupJournalAOSRTab()
        self.setupInformationTab()
        self.setupOtherTablesTab()
        
        self.tab_widget.currentChanged.connect(self.tab_changed)
        self.table.itemChanged.connect(lambda item, table=self.table: self.item_changed(item, table))

    def setupJournalAOSRTab(self):
        # Remove existing layout if it exists
        existing_layout = self.tabs['Журнал АОСР'].layout()
        if existing_layout is not None:
            QWidget().setLayout(existing_layout)
        layout = QVBoxLayout()
        self.tabs['Журнал АОСР'].setLayout(layout)

        # Determine the number of columns based on the region
        column_count = 15 if self.current_region == 'информация_рф' else 14
        self.table = QTableWidget(0, column_count)

        # Set the header labels based on the region
        headers = [
            'Номер\nАОСР', 'Город\nстроительства', 'Наименование\n объекта', 'Дата\nначала\nработ', 'Дата\nокончания\nработ',
            'Дата подписания\nакта', '1. К\nосвидетельствованию\nпредъявлены\nработы', 'Вид и\nобъём\nработ\n(выгрузка ведомости)',
            'Работы выполнены\nпо\nпроектно-сметной\nдок.', 'При\nвыполнении работ\nприменены', 'При\nвыполнении работ\nотклонения',
            'Разрешается\nдля\nпроизводства работ', 'Сформировать\nакт', 'Исполнительная\nсхема'
        ]

        # Add the additional column header if the current region is РФ
        if self.current_region == 'информация_рф':
            headers.append('Работа\nвыполнена\nв соответствии\nс НТД')

        self.table.setHorizontalHeaderLabels(headers)
        self.table.resizeRowsToContents()
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)
        header.setSectionResizeMode(12, QHeaderView.ResizeToContents)  # Adjust the "Сформировать акт" column width

        self.table.itemChanged.connect(self.capture_change)
        self.initialize_default_values()
        for column in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(column)
            if header_item:  # Ensure header_item is not None
                header_item.setToolTip(header_item.text().replace('\n', ' '))

        button_panel = QWidget()
        button_layout = QHBoxLayout()
        button_panel.setLayout(button_layout)
        self.table.setColumnHidden(1, True)
        self.table.setColumnHidden(2, True)

        btn_create_act = QPushButton('Создать акт')
        btn_create_act.clicked.connect(self.export_to_pdf_and_xls)
        btn_select_volume = QPushButton('Выбрать объём')
        btn_select_volume.clicked.connect(self.open_volume_selection)
        btn_select_mtr = QPushButton('Выбрать МТР')
        btn_select_mtr.clicked.connect(self.open_mtr_selection)
        btn_agreements = QPushButton('Выбрать согласования')
        btn_agreements.clicked.connect(self.open_agreement_selection)
        btn_select_scheme = QPushButton('Выбрать схему')
        btn_select_scheme.clicked.connect(self.open_scheme_selection)
        btn_clear_cell = QPushButton('Очистить ячейку')
        btn_clear_cell.clicked.connect(self.clear_selected_cell)
        upload_button = QPushButton(QIcon('images/import.jpeg'), '')
        upload_button.clicked.connect(self.upload_excel_data)

        download_button = QPushButton(QIcon('images/export.jpeg'), '')
        download_button.clicked.connect(self.download_csv_data)
        btn_undo = QPushButton(QIcon('images/reverse.jpeg'), '')
        btn_undo.clicked.connect(self.undo_stack.undo)

        new_button = QPushButton(QIcon('images/new.jpeg'), '')
        new_button.clicked.connect(self.new_project)

        

        button_layout.addWidget(new_button)
        button_layout.addWidget(btn_create_act)
        button_layout.addWidget(btn_clear_cell)
        if self.current_region == 'информация_рф':
            btn_select_ntd = QPushButton('Выбрать НТД')
            btn_select_ntd.clicked.connect(self.open_ntd_selection)
            button_layout.addWidget(btn_select_ntd)
        button_layout.addWidget(btn_select_mtr)
        button_layout.addWidget(btn_select_volume)
        button_layout.addWidget(btn_agreements)
        button_layout.addWidget(btn_select_scheme)
        button_layout.addWidget(upload_button)
        button_layout.addWidget(download_button)
        button_layout.addWidget(btn_undo)
        layout.addWidget(button_panel)
        layout.addWidget(self.table)

        self.table.itemChanged.connect(lambda item, table=self.table: self.item_changed(item, table))
        btn_add = QPushButton('Добавить запись')
        btn_remove = QPushButton('Удалить запись')
        btn_add.clicked.connect(self.add_record)
        btn_remove.clicked.connect(self.remove_record)
        layout.addWidget(btn_add)
        layout.addWidget(btn_remove)
        self.main_table = self.table



    def setupInformationTab(self):
        layout = QVBoxLayout()
        self.tabs['Информация'].setLayout(layout)

        self.chkRF = QCheckBox("РФ")
        self.chkKZ = QCheckBox("КЗ")
        self.chkKZ.setChecked(True)
        self.chkRF.toggled.connect(lambda: self.onRegionChanged('информация_рф'))
        self.chkKZ.toggled.connect(lambda: self.onRegionChanged('информация_кз'))
        layout.addWidget(self.chkRF)
        layout.addWidget(self.chkKZ)

        information_table = QTableWidget(0, 1)

        self.other_tables['Информация'] = information_table
        layout.addWidget(information_table)

    def setupOtherTables(self):
        tab_configs = {
            'Исп. схемы': ['Наименование схем', '№', ''],
            'Согласования': ['Наименование документа'],
            'Виды и объемы работ': ['', 'Ед.изм', 'Номер'],
            'Реестр ИД': ['№ акта', 'Наименование', 'Кол-во листов', 'Примечание'],
            'Ведомость МТР': ['Объект контроля', 'Сертификаты, паспорта и иные документы', '*Акты входного контроля'],
            'ВОР': ['Наименование выполненных работ', 'Ед.изм', 'Кол-во', 'Примечание'],
        }
        for name, headers in tab_configs.items():
            table = QTableWidget(0, len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            table.horizontalHeader().setStretchLastSection(True)
            self.initialize_default_values()
            self.other_tables[name] = table

    def setupOtherTablesTab(self):
        layout = QVBoxLayout()
        other_tables_tab = QWidget()
        other_tables_tab.setLayout(layout)
        self.tab_widget.addTab(other_tables_tab, QIcon('images/1.png'), 'База данных')

        self.menu_bar = QComboBox()
        self.menu_bar.addItems([
            'Исп. схемы', 'Согласования', 'Ведомость МТР', 'ВОР', 'Виды и объемы работ', 'Реестр ИД'
        ])

        self.menu_bar.currentTextChanged.connect(self.display_table)
        layout.addWidget(self.menu_bar)

        self.other_table_widget = QTableWidget()
        layout.addWidget(self.other_table_widget)

        self.setupOtherTables()

        button_panel = QWidget()
        button_layout = QHBoxLayout()
        button_panel.setLayout(button_layout)

        btn_add = QPushButton('Добавить запись')
        btn_remove = QPushButton('Удалить запись')
        btn_add.clicked.connect(self.add_other_record)
        btn_remove.clicked.connect(self.remove_other_record)

        button_layout.addWidget(btn_add)
        button_layout.addWidget(btn_remove)

        btn_save = QPushButton('Сохранить записи')
        btn_save.clicked.connect(self.save_table_data)
        button_layout.addWidget(btn_save)

        layout.addWidget(button_panel)

        self.load_table_data()
        self.display_table('Исп. схемы')


    

    def display_table(self, table_name):
        self.other_table_widget.setColumnCount(0)
        self.other_table_widget.setRowCount(0)
        if table_name in self.other_tables:
            table = self.other_tables[table_name]
            self.other_table_widget.setColumnCount(table.columnCount())
            self.other_table_widget.setHorizontalHeaderLabels([table.horizontalHeaderItem(i).text() for i in range(table.columnCount())])
            for row in range(table.rowCount()):
                self.other_table_widget.insertRow(row)
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        new_item = QTableWidgetItem(item.text())
                        new_item.setFlags(item.flags())
                        self.other_table_widget.setItem(row, col, new_item)
            self.other_table_widget.resizeColumnsToContents()  # Resize columns to fit contents

    
    def new_project(self):
        try:
            
            folder_path = 'docs/'
            files = os.listdir(folder_path)
            csv_files = [f for f in files if f.endswith('.csv')]
            for file_name in csv_files:
                with open(f'{folder_path}{file_name}', 'w', newline='', encoding='utf-8') as file:
                    pass

            
            for table_name, table in self.other_tables.items():
                table.setRowCount(0)  

            self.table.setRowCount(0)  

            QMessageBox.information(self, 'Успех', 'Проект успешно создан.')
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', f'Не удалось создать проект: {str(e)}')
    
    def capture_change(self, item):
        
        old_value = item.data(QtCore.Qt.UserRole)  
        new_value = item.text()
        if item.column() == 7 and old_value != new_value:
            self.export_work_volume_to_general_ledger()
        if old_value is None or old_value != new_value:
            command = EditCellCommand(self.table, item.row(), item.column(), old_value, new_value)
            self.undo_stack.push(command)
        item.setData(QtCore.Qt.UserRole, new_value)  



    def clear_selected_cell(self):
        selected_items = self.table.selectedItems()
        if selected_items:
            for item in selected_items:
                row, col = item.row(), item.column()
                
                if self.table.item(row, col) is not None:
                    self.table.item(row, col).setText('')  
                    if col == 7:
                        self.export_work_volume_to_general_ledger()

    
    def export_registry_to_xls_and_pdf(self):
        try:
            
            xls_path = 'Acts/XLS/Реестр_ИД.xlsx'
            pdf_path = 'Acts/PDF/Реестр_ИД.pdf'
            os.makedirs('Acts/XLS', exist_ok=True)
            os.makedirs('Acts/PDF', exist_ok=True)

            
            if os.path.exists(xls_path):
                wb = openpyxl.load_workbook(xls_path)
                ws = wb.active
                ws.delete_rows(2, ws.max_row)  
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(['№п/п', '№ акта', 'Наименование', 'Кол-во листов', 'Примечание'])  

            
            ws.title = "Реестр ИД"

            
            table = self.other_tables['Реестр ИД']
            row_count = 1  

            
            for row_index in range(table.rowCount()):
                row_data = [table.item(row_index, col).text() if table.item(row_index, col) else '' for col in range(table.columnCount())]
                row_data.insert(0, str(row_count))  
                ws.append(row_data)
                row_count += 1

            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 72
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            ws['C1'].alignment = Alignment(horizontal='center')

            
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

            
            wb.save(xls_path)
            wb.close()

            xlApp = client.Dispatch("Excel.Application")
            try:
                            books = xlApp.Workbooks.Open(os.path.abspath(xls_path))
                            ws = books.Worksheets[0]  
                            ws.Visible = 1
                            ws.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            finally:
                            books.Close()
                            xlApp.Quit()  
                            del xlApp

            QMessageBox.information(self, 'Успех', 'Реестр ИД успешно экспортирован в XLS и PDF.')
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', f'Не удалось экспортировать Реестр ИД: {str(e)}')
    
    def export_vor_to_xls_and_pdf(self):
        try:
            
            xls_path = 'Acts/XLS/Ведомость_объёмов_работ.xlsx'
            pdf_path = 'Acts/PDF/Ведомость_объёмов_работ.pdf'
            os.makedirs('Acts/XLS', exist_ok=True)
            os.makedirs('Acts/PDF', exist_ok=True)

            
            if os.path.exists(xls_path):
                wb = openpyxl.load_workbook(xls_path)
                ws = wb.active
                ws.delete_rows(2, ws.max_row)  
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(['№п/п', 'Наименование выполненных работ', 'Ед. изм', 'Кол-во', 'Примечание'])  

            
            ws.title = "ВОР"

            
            table = self.other_tables['ВОР']
            row_count = 1  

            
            for row_index in range(table.rowCount()):
                row_data = [table.item(row_index, col).text() if table.item(row_index, col) else '' for col in range(table.columnCount())]
                row_data.insert(0, str(row_count))  
                ws.append(row_data)
                row_count += 1

            
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 72
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 15

            ws['B1'].alignment = Alignment(horizontal='center')

            
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0

            
            wb.save(xls_path)
            wb.close()

            
            xlApp = client.Dispatch("Excel.Application")
            try:
                            books = xlApp.Workbooks.Open(os.path.abspath(xls_path))
                            ws = books.Worksheets[0]  
                            ws.Visible = 1
                            ws.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            finally:
                            books.Close()
                            xlApp.Quit()  
                            del xlApp

            QMessageBox.information(self, 'Успех', 'ВОР успешно экспортирован в XLS и PDF.')
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', f'Не удалось экспортировать ВОР: {str(e)}')




    def open_mtr_selection(self):
        dialog = MTRSelectionDialog(self)
        dialog.exec_()
    def setupOtherTabs(self):
        self.other_tables = {}
        tab_configs = {
            'Исп. схемы': ['Наименование схем', '№', ''],
            'Согласования': ['Наименование документа'],
            'Информация' : [''],
            'Виды и объемы работ': ['', 'Ед.изм', 'Номер'],
            'Реестр ИД': ['№ акта', 'Наименование', 'Кол-во листов', 'Примечание'],
            'Ведомость МТР': ['Объект контроля', 'Сертификаты, паспорта и иные документы', '*Акты входного контроля'],
            'ВОР': ['Наименование выполненных работ', 'Ед.изм', 'Кол-во', 'Примечание'],
        }
        for name, headers in tab_configs.items():
                tab = self.tabs[name]
                layout = QVBoxLayout()
                table = QTableWidget(0, len(headers))
                table.resizeRowsToContents()  
                table.setHorizontalHeaderLabels(headers)
                table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                table.horizontalHeader().setStretchLastSection(True)
                self.initialize_default_values()
                for column in range(table.columnCount()):
                    header_item = table.horizontalHeaderItem(column)
                    header_item.setToolTip(header_item.text())  

                tab.setLayout(layout)
                layout.addWidget(table)
                
                if name == 'Информация':
                    self.chkRF = QCheckBox("РФ")
                    self.chkKZ = QCheckBox("КЗ")
                    self.chkKZ.setChecked(True)
                    self.chkRF.toggled.connect(lambda: self.onRegionChanged('информация_рф'))
                    self.chkKZ.toggled.connect(lambda: self.onRegionChanged('информация_кз'))
                    layout.addWidget(self.chkRF)
                    layout.addWidget(self.chkKZ)
                button_panel = QWidget()
                button_layout = QHBoxLayout()
                button_panel.setLayout(button_layout)

                btn_add = QPushButton('Добавить запись')
                btn_remove = QPushButton('Удалить запись')
                
                btn_add.clicked.connect(lambda checked, t=table: self.add_other_record(t))
                btn_remove.clicked.connect(lambda checked, t=table: self.remove_other_record(t))
                
                button_layout.addWidget(btn_add)
                button_layout.addWidget(btn_remove)

                if name == "Реестр ИД":
                    btn_refresh_reg = QPushButton('Обновить реестр')
                    btn_refresh_reg.clicked.connect(self.reload_reg)
                    button_layout.addWidget(btn_refresh_reg)
                    btn_export_reg = QPushButton('Экспорт реестра')
                    btn_export_reg.clicked.connect(self.export_registry_to_xls_and_pdf)
                    button_layout.addWidget(btn_export_reg)
                if name == 'ВОР':
                    btn_refresh_reg = QPushButton('Обновить ведомость')
                    btn_refresh_reg.clicked.connect(self.reload_ov)
                    button_layout.addWidget(btn_refresh_reg)
                    btn_export_reg = QPushButton('Экспорт ведомости')
                    btn_export_reg.clicked.connect(self.export_vor_to_xls_and_pdf)
                    button_layout.addWidget(btn_export_reg)
                else:
                    btn_save = QPushButton('Сохранить записи')
                    button_layout.addWidget(btn_save)
                    btn_save.clicked.connect(self.save_table_data)
                layout.addWidget(button_panel)

                self.other_tables[name] = table

    def onRegionChanged(self, region):
        if region == 'информация_рф' and self.chkRF.isChecked():
            self.chkKZ.setChecked(False)
            self.current_region = 'информация_рф'
        elif region == 'информация_кз' and self.chkKZ.isChecked():
            self.chkRF.setChecked(False)
            self.current_region = 'информация_кз'
        self.load_information_data()
        self.setupJournalAOSRTab()
        with open('docs/журнал_аоср.csv', 'r', encoding='utf-8') as file:
                self.table.setRowCount(0)
                reader = csv.reader(file)
                for row_index, row_data in enumerate(reader):
                    self.table.insertRow(row_index)
                    self.set_default_value(row_index)
                    for col_index, value in enumerate(row_data):
                        if "Unnamed" in value:
                            value = ""
                        if col_index in [3, 4, 5]:
                            date = QDate.fromString(value, "MM/dd/yyyy") if value else QDate.currentDate()
                            date_editor = WheelIgnoredDateEdit(self)
                            date_editor.setDate(date)
                            date_editor.setProperty('oldDate', date_editor.date())
                            date_editor.setCalendarPopup(True)
                            date_editor.dateChanged.connect(lambda new_date, editor=date_editor, row=row_index, col=col_index: self.create_date_changed_handler(editor, row, col)(new_date))
                            self.table.setCellWidget(row_index, col_index, date_editor)
                        if col_index == 12:
                            checkbox_11 = QCheckBox()
                            checkbox_11.setChecked(row_data[12] == 'True')
                            checkbox_11.setProperty('oldValue', checkbox_11.isChecked())
                            checkbox_11.toggled.connect(lambda state, box=checkbox_11: self.update_registry_on_change(box, state))
                            self.table.setCellWidget(row_index, col_index, checkbox_11)
                        else:
                            item = NumericTableWidgetItem(value if value else '')
                            item.setTextAlignment(Qt.AlignCenter)  # Align text to center
                            if col_index == 9 and not value:
                                item = NumericTableWidgetItem('Не применялась')
                                item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                            if col_index == 10 and not value:
                                item = NumericTableWidgetItem('Отсутствуют')
                                item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                            self.table.setItem(row_index, col_index, item)
                            self.table.resizeRowToContents(item.row())
                            item.setToolTip(value)
                            item.setTextAlignment(Qt.AlignCenter)  # Align text to center
        self.load_and_display_excel_data()

    def add_other_record(self, table):
        row_count = table.rowCount()
        table.insertRow(row_count)
        for col_index in range(table.columnCount()):
            item = QTableWidgetItem('')
            table.setItem(row_count, col_index, item)

    def remove_other_record(self, table):
        current_row = table.currentRow()
        if current_row != -1:
            table.removeRow(current_row)
        else:
            QMessageBox.warning(self, 'Ошибка', 'Выберите строку для удаления')
    
    def open_agreement_selection(self):
        dialog = AgreementSelectionDialog(self)
        dialog.exec_()

    def open_scheme_selection(self):
        dialog = SchemeSelectionDialog(self)
        dialog.exec_()

    def initialize_default_values(self):
        for row in range(self.table.rowCount()):
            self.set_default_value(row)


    def upload_excel_data(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Открыть файл", "", "Excel Files (*.xlsx)")
        if file_path:
            try:
                wb = load_workbook(file_path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    data = pd.DataFrame(sheet.values)
                    headers = data.iloc[0]
                    data = data[1:]
                    data.columns = headers
                    data = data.reindex(columns=headers)  

                    csv_path = f"docs/{sheet_name}.csv"
                    data.to_csv(csv_path, index=False, header=True, encoding='utf-8')
                self.load_table_data()
                QMessageBox.information(self, 'Успех', 'Данные загружены.')
            except Exception as e:
                QMessageBox.warning(self, 'Ошибка', f'Не удалось загрузить данные: {str(e)}')

    def download_csv_data(self):
        self.save_changes()
        folder_path = 'docs/'
        files = os.listdir(folder_path)
        csv_files = [f for f in files if f.endswith('.csv')]
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for file_name in csv_files:
            try:
                data = pd.read_csv(f'{folder_path}{file_name}', keep_default_na=False)  
                ws = wb.create_sheet(title=file_name.replace('.csv', ''))

                for col_num, col_name in enumerate(data.columns, start=1):
                    ws.cell(row=1, column=col_num, value=col_name)
                
                for row_num, row in enumerate(data.itertuples(index=False, name=None), start=2):
                    for col_num, item in enumerate(row, start=1):
                        ws.cell(row=row_num, column=col_num, value=item)
        
            except Exception as e:
                pass
            
        wb.save('docs/project.xlsx')
        QMessageBox.information(self, 'Успех', 'Все данные выгружены в Excel файл project.xlsx.')


    def set_default_value(self, row_index):
        item_8 = QTableWidgetItem('Не применялась')
        item_9 = QTableWidgetItem('Отсутствуют')
        checkbox_11 = QCheckBox()
        checkbox_11.setChecked(False)
        checkbox_11.setProperty('oldValue', checkbox_11.isChecked())
        
        item_8.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
        item_9.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)

        self.table.setItem(row_index, 9, item_8)
        self.table.setItem(row_index, 10, item_9)
        self.table.setCellWidget(row_index, 12, checkbox_11)
        checkbox_11.toggled.connect(lambda state, box=checkbox_11: self.update_registry_on_change(box, state))

        if self.current_region == 'информация_рф':
            item_ntd = QTableWidgetItem('')
            item_ntd.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_index, 14, item_ntd)

    def open_ntd_selection(self):
        selected_row = self.table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, 'Ошибка', 'Не выбрана строка для добавления НТД.')
            return

    def update_registry_on_change(self, checkbox, new_value):
        try:
            with open('docs/реестр_ид.csv', 'w+', newline='', encoding='utf-8') as reg_file:
                reg_writer = csv.writer(reg_file, quoting=csv.QUOTE_ALL)
                for row in range(self.table.rowCount()):
                    checkbox = self.table.cellWidget(row, 12)
                    if checkbox and checkbox.isChecked():
                        aosr_number = self.table.item(row, 0).text() if self.table.item(row, 0) else ''
                        column_6_value = self.table.item(row, 6).text() if self.table.item(row, 6) else ''
                        column_13_value = self.table.item(row, 13).text().split('; ') if self.table.item(row, 13) else []

                        reg_writer.writerow([f"АОСР № {aosr_number}", column_6_value, None, None])
                        for item in column_13_value:
                            reg_writer.writerow(['', item, None, None])

                        selected_mtrs = self.get_selected_mtr_data(row)
                        for mtr in selected_mtrs:
                            reg_writer.writerow(['', mtr[2], None, None])

            self.export_work_volume_to_general_ledger()

            old_value = checkbox.property('oldValue') or checkbox.isChecked()
            if old_value != new_value:
                command = EditComboCommand(checkbox, old_value, new_value)
                self.undo_stack.push(command)
            checkbox.setProperty('oldValue', new_value)
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', f'Не удалось обновить реестр: {e}')


    def export_to_pdf_and_xls(self):
        self.save_changes()
        journal_path = 'docs/журнал_аоср.csv'
        info_path = f'docs/{self.current_region}.csv'
        template_path = 'docs/blank_kz.xlsx' if self.current_region == 'информация_кз' else 'docs/blank_ru.xlsx'
        registry_path = 'docs/реестр_ид.csv'
        xls_output_dir = 'Acts/XLS'
        pdf_output_dir = 'Acts/PDF'
        default_font = Font(name='Times New Roman', size=11)
            
        if os.path.exists(xls_output_dir):
            shutil.rmtree(xls_output_dir)
        os.makedirs(xls_output_dir, exist_ok=True)
        if os.path.exists(pdf_output_dir):
            shutil.rmtree(pdf_output_dir)
        os.makedirs(pdf_output_dir, exist_ok=True)
        
        try:
            journal_df = pd.read_csv(journal_path, header=None)
            info_df = pd.read_csv(info_path, header=None)
            
            with open(registry_path, 'w+', newline='', encoding='utf-8') as reg_file:
                reg_writer = csv.writer(reg_file, quoting=csv.QUOTE_ALL)
                
                for row in range(self.table.rowCount()):
                    if self.table.cellWidget(row, 12) and self.table.cellWidget(row, 12).currentText() == 'Да':
                        aosr_number = self.table.item(row, 0).text()
                        column_6_value = self.table.item(row, 6).text() if self.table.item(row, 6) else ''
                        column_13_value = self.table.item(row, 13).text().split('; ') if self.table.item(row, 13) else ''
                        reg_writer.writerow([f"АОСР № {aosr_number}", column_6_value, None, None])
                        for itemss in column_13_value:
                            reg_writer.writerow(['', itemss, None, None])
                        selected_mtrs = self.get_selected_mtr_data(row)
                        for mtr in selected_mtrs:
                            reg_writer.writerow(['', mtr[2], None, None])
                        
                        xls_filename = f"{xls_output_dir}/{aosr_number}.xlsx"
                        pdf_filename = f"{pdf_output_dir}/{aosr_number}.pdf"
                    
                        with open(template_path, "rb") as f_template, open(xls_filename, "wb") as f_output:
                            f_output.write(f_template.read())
                        
                        wb = load_workbook(xls_filename)
                        sheet = wb.active
                        cell_mapping = ['F3', 'A6', 'A8', 'D41', 'D42', 'H6', 'A29', None, 'A32', 'A35', 'A39', 'A47', None, None] if self.current_region == 'информация_кз' else ['B29', None, 'A3', 'E84', 'E85', 'H29', 'A67', None, 'A71', 'A76', 'A88', 'A94', 'A102']
                        journal_row_data = journal_df.iloc[row].tolist()
                        for index, cell in enumerate(cell_mapping):
                            if cell:
                                cell_value = sheet[cell]
                                cell_value.font = default_font
                                date_cells = ['H6', 'D41', 'D42'] if self.current_region == 'информация_кз' else ['E84', 'E85', 'H29']
                                useful_cells = ['A35'] if self.current_region == 'информация_кз' else ['A71']
                                if cell in date_cells:
                                    date_str = journal_row_data[index]
                                    formatted_date = datetime.datetime.strptime(date_str, '%m/%d/%Y').strftime('%d.%m.%Y')
                                    sheet[cell] = formatted_date
                                elif cell in useful_cells:
                                    new_str = journal_row_data[index].split("; ")
                                    ddata = ''
                                    for item in new_str:
                                        ddata += item
                                        ddata += '\n'
                                    sheet[cell] = ddata.strip()
                                    sheet.row_dimensions[sheet[cell].row].height = 15 * len(new_str)
                                elif cell == 'A102':
                                    ddata = ''
                                    for item in journal_row_data[9].split("; "):
                                        last_str = item.split(' - ')
                                        if len(last_str) == 2:
                                            ddata += last_str[1]
                                            ddata += '\n'
                                    new_str = journal_row_data[-1].split("; ")
                                    for item in new_str:
                                        ddata += item
                                        ddata += '\n'
                                    sheet[cell] = ddata.strip()
                                    sheet.row_dimensions[sheet[cell].row].height = 15 * (len(new_str) + len(journal_row_data[9].split("; ")))
                                elif cell == 'A88':
                                    ddata = ''
                                    new_str = journal_row_data[-1].split("; ")
                                    for item in new_str:
                                        ddata += item
                                        ddata += '\n'
                                    sheet[cell] = ddata.strip()
                                    sheet['A71'] = ddata.strip()
                                    sheet.row_dimensions[sheet['A71'].row].height = 15 * len(new_str)
                                    sheet['A80'] = ddata.strip()
                                    sheet.row_dimensions[sheet['A80'].row].height = 15 * len(new_str)
                                    sheet.row_dimensions[sheet[cell].row].height = 15 * len(new_str)
                                else:
                                    sheet[cell] = journal_row_data[index]
                                sheet[cell].alignment = Alignment(wrapText=True)
                                sheet[cell].font = default_font
                        
                        info_mapping = {'A13': 4, 'A16': 7, 'A19': 10, 'A23': 14, 'A25': 16, 'H51': (22, 1), 'H55': (26, 1), 'H58': (29, 1), 'H61': (32, 1)} if self.current_region == 'информация_кз' else {'A6': 6, 'A8': 7, 'A10': 8, 'A13': 11, 'A15': 12, 'A17': 13, 'A20': 16, 'A22': 17, 'A24': 18, 'A33': 21, 'A35': 22, 'A37': 23, 'A40': 26, 'A44': 29, 'A46': 30, 'A49': 33, 'A51': 34, 'A53': 35, 'A56': 38, 'A58': 39, 'A63': 41, 'C99': (43, 1), 'B109': (47, 1), 'B114': (50, 1), 'B120': (54, 1), 'B127': (57, 1), 'B134': (60, 1)}
                        for cell, ref in info_mapping.items():
                            if isinstance(ref, tuple):
                                sheet[cell] = info_df.iloc[ref[0], ref[1]]
                            else:
                                sheet[cell] = info_df.iloc[ref, 0]
                        
                        wb.save(xls_filename)
                        wb.close()
                        
                        xlApp = client.Dispatch("Excel.Application")
                        try:
                            books = xlApp.Workbooks.Open(os.path.abspath(xls_filename))
                            ws = books.Worksheets[0]  
                            ws.Visible = 1
                            ws.ExportAsFixedFormat(0, os.path.abspath(pdf_filename))
                        finally:
                            books.Close()
                            xlApp.Quit() 
                            del xlApp
                        gc.collect()
            
            self.reload_reg()

            QMessageBox.information(self, 'Успешно', f'Акты сформированы')

        except Exception as e:
           QMessageBox.warning(self, 'Ошибка', f'Не удалось сформировать акт {e}')
            
                
    def get_selected_mtr_data(self, aosr_row):
        mtr_table = self.other_tables['Ведомость МТР']
        selected_mtrs = []
        

        selected_ids = self.table.item(aosr_row, 9).text().split('; ') if self.table.item(aosr_row, 9) else []
        for mtr_id in selected_ids:
            for row in range(mtr_table.rowCount()):
                mtr_number = mtr_table.item(row, 1).text() if mtr_table.item(row, 1) else ''
                if mtr_number in mtr_id:
                    scheme = mtr_table.item(row, 0).text() if mtr_table.item(row, 0) else ''
                    quality_docs = mtr_table.item(row, 2).text() if mtr_table.item(row, 2) else ''
                    selected_mtrs.append((scheme, mtr_number, quality_docs))
        return selected_mtrs

    def create_xlsx_from_template(self, registry_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry Data"
        
        headers = ['АОСР №', 'Наименование объекта', 'Документы и сертификаты о качестве', 'Исполнительные схемы']
        ws.append(headers)
        
        with open(registry_path, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                ws.append(row)
        
        output_xlsx_path = 'docs/реестр_ид.xlsx'
        wb.save(output_xlsx_path)
        wb.close()
    
    def tab_changed(self, index):
        tab_name = self.tab_widget.tabText(index)
        self.setWindowTitle(f'АОСР - {tab_name}')

    def load_table_data(self):
        try:
            self.table.itemChanged.disconnect(self.item_changed)
        except TypeError:
            pass
        try:
            with open('docs/журнал_аоср.csv', 'r', encoding='utf-8') as file:
                self.table.setRowCount(0)
                reader = csv.reader(file)
                for row_index, row_data in enumerate(reader):
                    self.table.insertRow(row_index)
                    self.set_default_value(row_index)
                    for col_index, value in enumerate(row_data):
                        if "Unnamed" in value:
                            value = ""
                        if col_index in [3, 4, 5]:
                            date = QDate.fromString(value, "MM/dd/yyyy") if value else QDate.currentDate()
                            date_editor = WheelIgnoredDateEdit(self)
                            date_editor.setDate(date)
                            date_editor.setProperty('oldDate', date_editor.date())
                            date_editor.setCalendarPopup(True)
                            date_editor.dateChanged.connect(lambda new_date, editor=date_editor, row=row_index, col=col_index: self.create_date_changed_handler(editor, row, col)(new_date))
                            self.table.setCellWidget(row_index, col_index, date_editor)
                        if col_index == 12:
                            checkbox_11 = QCheckBox()
                            checkbox_11.setChecked(row_data[12] == 'True')
                            checkbox_11.setProperty('oldValue', checkbox_11.isChecked())
                            checkbox_11.toggled.connect(lambda state, box=checkbox_11: self.update_registry_on_change(box, state))
                            self.table.setCellWidget(row_index, col_index, checkbox_11)
                        else:
                            item = NumericTableWidgetItem(value if value else '')
                            item.setTextAlignment(Qt.AlignCenter)  # Align text to center
                            if col_index == 9 and not value:
                                item = NumericTableWidgetItem('Не применялась')
                                item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                            if col_index == 10 and not value:
                                item = NumericTableWidgetItem('Отсутствуют')
                                item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                            self.table.setItem(row_index, col_index, item)
                            self.table.resizeRowToContents(item.row())
                            item.setToolTip(value)
                            item.setTextAlignment(Qt.AlignCenter)  # Align text to center
            for tab_name, table in self.other_tables.items():
                if tab_name == 'Информация':
                    self.load_and_display_excel_data()
                    self.load_information_data()
                else:
                    file_path = f'docs/{tab_name.lower().replace(" ", "_")}.csv'
                    with open(file_path, 'r', encoding='utf-8') as file:
                        table.setRowCount(0)
                        reader = csv.reader(file)
                        for row_data in reader:
                            if any(row_data): 
                                row_index = table.rowCount()
                                table.insertRow(row_index)
                                for col_index, value in enumerate(row_data):
                                    if "Unnamed:" in value:
                                        value = ""
                                    item = NumericTableWidgetItem(value if value else '')
                                    item.setToolTip(value)
                                    item.setTextAlignment(Qt.AlignCenter)  # Align text to center
                                    table.setItem(row_index, col_index, item)
                                    table.resizeRowToContents(item.row())
        except FileNotFoundError:
            QMessageBox.warning(self, 'Ошибка', 'Файл данных не найден.')
        finally:
            self.table.itemChanged.connect(lambda item, table=self.table: self.item_changed(item, table))
            for _, tables in self.other_tables.items():
                tables.itemChanged.connect(lambda item, table=tables: self.item_changed(item, table))
            self.table.sortItems(0, QtCore.Qt.AscendingOrder)
            self.validate_all_dates()


    def load_information_data(self):
        """ Load the first two rows from информация.csv and set them as values for the 2nd and 3rd columns in the AOSR journal table. """
        csv_path = f'docs/{self.current_region}.csv'
        try:
            with open(csv_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                info_data = list(reader)[:4 if self.current_region == 'информация_кз' else 2]  

                for row_index in range(self.table.rowCount()):
                    if self.table.rowCount() < row_index + 1:
                        self.table.insertRow(self.table.rowCount())  

                    city_data = info_data[0][0]
                    self.table.setItem(row_index, 1, QTableWidgetItem(city_data))

                    object_name_data = info_data[3][0] if self.current_region == 'информация_кз' else info_data[1][0]
                    self.table.setItem(row_index, 2, QTableWidgetItem(object_name_data))

        except Exception:
            pass

    def reload_ov(self):
        with open('docs/вор.csv', 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            table = self.other_tables['ВОР']
            table.setRowCount(0)
            for row_index, row_data in enumerate(reader):
                    if any(row_data):  
                        table.insertRow(row_index)
                        for col_index, value in enumerate(row_data):
                            item = NumericTableWidgetItem(value if value else '')
                            table.setItem(row_index, col_index, item)
                            item.setToolTip(value)
                            table.resizeRowToContents(item.row())

    def reload_reg(self):
        with open('docs/реестр_ид.csv', 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            table = self.other_tables['Реестр ИД']
            table.setRowCount(0)
            for row_data in reader:
                if any(row_data):  
                    row_index = table.rowCount()
                    table.insertRow(row_index)
                    for col_index, value in enumerate(row_data):
                        item = NumericTableWidgetItem(value if value else '')
                        table.setItem(row_index, col_index, item)
                        item.setToolTip(value)
                        table.resizeRowToContents(item.row())

    def create_date_changed_handler(self, editor, row, col):
        def handle_date_changed(new_date):
            old_date = editor.property('oldDate') if editor.property('oldDate') is not None else editor.date()
            if old_date != new_date:
                command = EditDateCommand(editor, old_date, new_date)
                self.undo_stack.push(command)
            editor.setProperty('oldDate', new_date)  
            self.save_changes()  
        return handle_date_changed


    def load_and_display_excel_data(self):
        csv_path = f'docs/{self.current_region}.csv'
        try:
            table = self.other_tables['Информация']
            table.setRowCount(0)
            skip_rows = 2 if self.current_region == 'информация_рф' else 0
            data_frame = pd.read_csv(csv_path, header=None, skiprows=skip_rows)
            data_frame = data_frame.fillna('')

            table.setRowCount(len(data_frame.index))
            table.setColumnCount(len(data_frame.columns))
            table.horizontalHeader().setVisible(False)  
            editable_rows_first_column = {1, 2, 5, 8, 11, 15, 17} if self.current_region == 'информация_кз' else {2, 5, 6, 7, 10, 11, 12, 15, 16, 17, 20, 21, 22, 25, 28, 29, 32, 33, 34, 37, 38, 41}
            
            for index, row in data_frame.iterrows():
                for col_index, value in enumerate(row):
                    if "Unnamed:" in value:
                        value = ""
                    item = QTableWidgetItem(str(value))
                    item.setFlags(Qt.ItemIsEnabled)  
                    if col_index == len(row) - 1 or (col_index == 0 and (index + 1) in editable_rows_first_column):
                        item.setFlags(Qt.ItemIsEditable | Qt.ItemIsEnabled)
                        if self.current_region == 'информация_кз':
                            if index + 1 in [5,8,11,15] or col_index==len(row)-1:
                                item.setTextAlignment(Qt.AlignCenter)
                        item.setBackground(QColor(255, 255, 255))  
                    else:
                        item.setBackground(QColor(240, 240, 240))  
                    table.setItem(index, col_index, item)
            table.resizeColumnsToContents()
        except Exception as e:
            print(e)

    
    def date_item_changed(self, row, column, date):
        self.row_modified[row] = True
        self.validate_date(row, column)

    def item_changed(self, item, table):
        self.row_modified[item.row()] = True
        table.resizeRowToContents(item.row())
        self.update_tooltip(item)

    def export_work_volume_to_general_ledger(self):
        try:
            with open('docs/вор.csv', 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                for row in range(self.table.rowCount()):
                    checkbox = self.table.cellWidget(row, 12)
                    if checkbox and checkbox.isChecked():
                        work_volume_item = self.table.item(row, 7)
                        if work_volume_item:
                            work_volumes = work_volume_item.text().split("; ")
                            for item in work_volumes:
                                work_volume_data = item.split("_")
                                if len(work_volume_data) >= 4:
                                    writer.writerow([
                                        work_volume_data[1],  # Наименование выполненных работ
                                        work_volume_data[3],  # Объем работы
                                        work_volume_data[2],  # Единица измерения
                                        self.table.item(row, 0).text()  # Номер АОСР
                                    ])
                                else:
                                    writer.writerow(["", "", "", ""])  # Write empty row if data is incomplete
                        else:
                            writer.writerow(["", "", "", ""])  # Write empty row if item is None
                    else:
                        writer.writerow(["", "", "", ""])  # Write empty row if checkbox is not checked

            self.reload_ov()
        except Exception as e:
            QMessageBox.warning(self, 'Ошибка', f'Не удалось экспортировать объемы работ в общий журнал: {e}')


    def update_tooltip(self, item):
        item.setToolTip(item.text()) 

    def get_mtr_data(self):
    
        mtr_table = self.other_tables['Ведомость МТР']
        data = []
        for row in range(mtr_table.rowCount()):
            
            row_data = tuple(mtr_table.item(row, col).text() if mtr_table.item(row, col) else '' for col in range(mtr_table.columnCount()))
            data.append(row_data)
        return data

    def get_agreement_data(self):
        agreement_table = self.other_tables['Согласования']
        data = []
        for row in range(agreement_table.rowCount()):
            row_data = tuple(agreement_table.item(row, col).text() if agreement_table.item(row, col) else '' for col in range(agreement_table.columnCount()))
            data.append(row_data)
        return data
    
    def get_scheme_data(self):
    
        scheme_table = self.other_tables['Исп. схемы']
        data = []
        for row in range(scheme_table.rowCount()):
            
            row_data = tuple(scheme_table.item(row, col).text() if scheme_table.item(row, col) else '' for col in range(scheme_table.columnCount()))
            data.append(row_data)
        return data
    
    def add_record(self):
        row_count = self.table.rowCount()
        self.table.insertRow(row_count)
        if row_count > 0:
            
            last_item = self.table.item(row_count - 1, 0)
            if last_item:
                last_aosr_number = int(last_item.text())
            else:
                last_aosr_number = 0
            next_aosr_number = last_aosr_number + 1
        else:
            next_aosr_number = 1  
        
        
        aosr_number_item = NumericTableWidgetItem(str(next_aosr_number))
        self.table.setItem(row_count, 0, aosr_number_item)
        self.set_default_value(row_count)
        self.table.setItem(row_count, 0, aosr_number_item)
        self.set_default_value(row_count)
        for col_index in range(1, self.table.columnCount()):
            if col_index in [3, 4, 5]:  
                date_editor = WheelIgnoredDateEdit(self)
                date_editor.setCalendarPopup(True)
                date_editor.setDate(QDate.currentDate())
                date_editor.dateChanged.connect(lambda new_date, editor=date_editor, row=row_count, col=col_index: self.create_date_changed_handler(editor, row, col)(new_date))
                date_editor.setProperty('oldDate', date_editor.date())
                self.table.setCellWidget(row_count, col_index, date_editor)
            else:
                item = NumericTableWidgetItem('')
                if col_index == 9:
                    item = NumericTableWidgetItem('Не применялась')
                    item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                if col_index == 10:
                    item = NumericTableWidgetItem('Отсутствуют')
                    item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                self.table.setItem(row_count, col_index, item)
        self.row_modified[row_count] = True

    def remove_record(self):
        current_row = self.table.currentRow()
        if current_row != -1:
            self.table.removeRow(current_row)
            self.save_changes()
        else:
            QMessageBox.warning(self, 'Ошибка', 'Выберите строку для удаления')

    def validate_date(self, row, column):
        date_editor = self.table.cellWidget(row, column)
        if not date_editor:
            return True  
        date = date_editor.date()
        if column == 3:  
            end_date_editor = self.table.cellWidget(row, 4)
            if end_date_editor and date > end_date_editor.date():
                date_editor.setStyleSheet("background-color: #ffcccc;")
                end_date_editor.setStyleSheet("background-color: #ffcccc;")
                return False
        elif column == 4:  
            start_date_editor = self.table.cellWidget(row, 3)
            if start_date_editor and date < start_date_editor.date():
                date_editor.setStyleSheet("background-color: #ffcccc;")
                start_date_editor.setStyleSheet("background-color: #ffcccc;")
                return False
        elif column == 5:  
            start_date_editor = self.table.cellWidget(row, 3)
            end_date_editor = self.table.cellWidget(row, 4)
            if (start_date_editor and date < start_date_editor.date()) or (end_date_editor and date < end_date_editor.date()):
                date_editor.setStyleSheet("background-color: #ffcccc;")
                return False
        date_editor.setStyleSheet("")  
        return True

    def validate_all_dates(self):
        for row in range(self.table.rowCount()):
            for col in [3, 4, 5]:  
                self.validate_date(row, col)

    def save_changes(self, filename='docs/журнал_аоср.csv'):
        with open(filename, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL) 
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    if col in [3, 4, 5]:  
                        date_editor = self.table.cellWidget(row, col)
                        date = date_editor.date().toString("MM/dd/yyyy") if date_editor else ""
                        row_data.append(date)
                    elif col == 12:  
                        combo_box = self.table.cellWidget(row, col)
                        value = combo_box.currentText() if combo_box else 'Нет'
                        row_data.append(value)
                    else:
                        item = self.table.item(row, col)
                        text = item.text() if item else ""
                        row_data.append(text)
                writer.writerow(row_data)

    def save_table_data(self):
        name = self.windowTitle()[7:]
        table = self.other_tables[name]
        if name == 'Информация':
            name = self.current_region
        filename = f"docs/{name.lower().replace(' ', '_')}.csv"
        with open(filename, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file, quoting=csv.QUOTE_ALL) 
            for row in range(table.rowCount()):
                row_data = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    text = item.text() if item else ""
                    row_data.append(text)
                writer.writerow(row_data)
    
    def open_volume_selection(self):
        selected_row = self.table.currentRow()
        if selected_row == -1:
            QMessageBox.warning(self, 'Ошибка', 'Не выбрана строка для добавления объема работы.')
            return

        dialog = VolumeSelectionDialog(self)
        
        volume_data = self.get_volume_data()
        dialog.set_data(volume_data)
        
        if dialog.exec_() == QtWidgets.QDialog.Accepted and dialog.selected_data:
            
            self.reload_ov()
            self.table.setItem(selected_row, 7, QtWidgets.QTableWidgetItem(dialog.selected_data))
            self.save_changes()
            self.export_work_volume_to_general_ledger()  

    def get_volume_data(self):
        
        
        data = []
        volume_table = self.other_tables['Виды и объемы работ']
        for row in range(volume_table.rowCount()):
            row_data = []
            for col in range(volume_table.columnCount()):
                item = volume_table.item(row, col)
                row_data.append(item.text() if item else '')
            data.append(row_data)
        return data
    

def main():
    app = QApplication(sys.argv)
    ex = AOSRApp()
    ex.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
